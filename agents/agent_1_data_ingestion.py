import sqlite3
from typing import List, Dict, Optional
from pathlib import Path
from openpyxl import load_workbook
import json
from datetime import datetime

REQUIRED_COLUMNS = {"name", "address", "website", "industry", "mcc"}

ALLOWED_COLUMNS = {
    "name",
    "address",
    "website",
    "industry",
    "mcc",
    "additional_data",
}

class DataIngestionAgent:

    def __init__(self, db_path: str = "merchants.db"):
        self.db_path = db_path
        self._initialize_database()

    # ----------------------------------
    # DATABASE INITIALIZATION
    # ----------------------------------
    def _initialize_database(self) -> None:
        self._create_tables()

    def _create_tables(self) -> None:
        query = """
        CREATE TABLE IF NOT EXISTS merchants (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            merchant_name TEXT,
            address TEXT,
            website TEXT,
            industry TEXT,
            mcc TEXT,
            additional_data TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """
        self._execute_query(query)

    def _get_connection(self) -> sqlite3.Connection:
        return sqlite3.connect(self.db_path)

    def _execute_query(self, query: str, params: tuple = ()) -> None:
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute(query, params)
        conn.commit()
        conn.close()

    # ----------------------------------
    # SINGLE INSERT (DIRECT INPUT)
    # ----------------------------------
    def insert_merchant(
        self,
        name: str,
        address: Optional[str] = None,
        website: Optional[str] = None,
        industry: Optional[str] = None,
        mcc: Optional[str] = None,
        additional_data: Optional[Dict] = None,
    ) -> Dict:

        if not name:
            raise ValueError("Merchant name is required")

        merchant_name = self._normalize_text(name)
        address = self._normalize_text(address)
        website = self._normalize_url(website)
        industry = self._normalize_text(industry)
        mcc = self._normalize_mcc(mcc)
        additional = self._serialize_additional(additional_data)

        conn = self._get_connection()
        cursor = conn.cursor()

        try:
            cursor.execute(
                """
                INSERT INTO merchants (merchant_name, address, website, industry, mcc, additional_data)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (merchant_name, address, website, industry, mcc, additional),
            )
            conn.commit()
            merchant_id = cursor.lastrowid

        except Exception as e:
            conn.rollback()
            raise Exception(f"Error inserting merchant: {str(e)}")

        finally:
            conn.close()

        return {
            "merchant_id": merchant_id,
            "merchant_name": merchant_name,
            "address": address,
            "website": website,
            "industry": industry,
            "mcc": mcc,
            "additional_data": additional,
        }

    # ----------------------------------
    # BULK INGESTION (XLSX)
    # ----------------------------------
    def ingest_from_xlsx(self, file_path: str) -> List[Dict]:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(cell).strip().lower() for cell in rows[0]]

        missing_columns = REQUIRED_COLUMNS - set(headers)
        if missing_columns:
            raise ValueError(f"Missing required columns: {missing_columns}")

        valid_indices = [i for i, col in enumerate(headers) if col in ALLOWED_COLUMNS]
        filtered_headers = [headers[i] for i in valid_indices]

        records = []
        for row in rows[1:]:
            record = {
                filtered_headers[i]: row[valid_indices[i]]
                for i in range(len(valid_indices))
            }
            records.append(record)

        return self._insert_records(records)

    # ----------------------------------
    # CORE INSERT LOGIC
    # ----------------------------------
    def _insert_records(self, records: List[Dict]) -> List[Dict]:
        conn = self._get_connection()
        cursor = conn.cursor()

        inserted_records = []

        try:
            for record in records:
                self._validate_record(record)

                merchant_name = self._normalize_text(record.get("name"))
                address = self._normalize_text(record.get("address"))
                website = self._normalize_url(record.get("website"))
                industry = self._normalize_text(record.get("industry"))
                mcc = self._normalize_mcc(record.get("mcc"))
                additional = self._serialize_additional(record.get("additional_data"))

                cursor.execute(
                    """
                    INSERT INTO merchants (merchant_name, address, website, industry, mcc, additional_data)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (merchant_name, address, website, industry, mcc, additional),
                )

                merchant_id = cursor.lastrowid

                inserted_records.append(
                    {
                        "merchant_id": merchant_id,
                        "merchant_name": merchant_name,
                        "address": address,
                        "website": website,
                        "industry": industry,
                        "mcc": mcc,
                        "additional_data": additional,
                    }
                )

            conn.commit()

        except Exception as e:
            conn.rollback()
            raise Exception(f"Error inserting records: {str(e)}")

        finally:
            conn.close()

        return inserted_records

    # ----------------------------------
    # FETCH
    # ----------------------------------
    def fetch_all(self) -> List[Dict]:
        conn = self._get_connection()
        cursor = conn.cursor()

        try:
            cursor.execute(
                "SELECT id, merchant_name, address, website, industry, mcc, additional_data, created_at FROM merchants"
            )
            rows = cursor.fetchall()

            columns = [
                "id",
                "merchant_name",
                "address",
                "website",
                "industry",
                "mcc",
                "additional_data",
                "created_at",
            ]

            return [dict(zip(columns, row)) for row in rows]

        finally:
            conn.close()

    # ----------------------------------
    # VALIDATION
    # ----------------------------------
    def _validate_record(self, record: Dict) -> None:
        if not record.get("name"):
            raise ValueError("Merchant name is required")

    # ----------------------------------
    # HELPERS
    # ----------------------------------
    def _normalize_text(self, value: Optional[str]) -> Optional[str]:
        return str(value).strip().title() if value else None

    def _normalize_url(self, url: Optional[str]) -> Optional[str]:
        if not url:
            return None
        url = str(url).strip().lower()
        return url if url.startswith("http") else f"https://{url}"

    def _normalize_mcc(self, mcc: Optional[str]) -> Optional[str]:
        return str(mcc).strip() if mcc else None

    def _serialize_additional(self, data: Optional[Dict]) -> Optional[str]:
        return json.dumps(data) if data else None

    # ----------------------------------
    # CLEAR DATA
    # ----------------------------------
    def clear_all_data(self) -> None:
        self._execute_query("DELETE FROM merchants")

if __name__ == "__main__":
    agent = DataIngestionAgent()
    # Example usage:
    agent.insert_merchant(name="Test Merchant", address="123 Main St", website="www.test.com", industry="Retail", mcc="1234")
    # agent.clear_all_data()  # Clear existing data before ingestion
    # agent.ingest_from_xlsx("merchants.xlsx")
    # agent.insert_merchant(name="Another Merchant", address="456 Elm St", website="www.another.com", industry="Food", mcc="5678")
    print(agent.fetch_all())